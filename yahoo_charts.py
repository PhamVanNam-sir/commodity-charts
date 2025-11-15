import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from bokeh.plotting import figure
from bokeh.models import HoverTool
from bokeh.embed import file_html
from bokeh.resources import INLINE
from openpyxl import Workbook
import os
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage


# Set style cho matplotlib
plt.style.use('seaborn-v0_8-darkgrid')

# Mapping tên commodity
COMMODITY_NAMES = {
    'HRC=F': 'Hot Rolled Coil',
    'CL=F': 'Crude Oil (WTI)',
    'BZ=F': 'Brent Crude',
    'NG=F': 'Natural Gas',
    'RB=F': 'Gasoline',
    'HO=F': 'Heating Oil',
    'EH=F': 'Ethanol',
    'GC=F': 'Gold',
    'SI=F': 'Silver',
    'HG=F': 'Copper',
    'PL=F': 'Platinum',
    'PA=F': 'Palladium',
    'ALI=F': 'Aluminum',
    'DX=F': 'Dollar Index',
}

def calculate_returns(prices):
    """Tính các loại returns"""
    df = pd.DataFrame({'Close': prices})
    
    # Daily return
    df['Daily'] = df['Close'].pct_change() * 100
    
    # Weekly return (5 trading days)
    df['Weekly'] = df['Close'].pct_change(periods=5) * 100
    
    # Monthly return (21 trading days)
    df['Monthly'] = df['Close'].pct_change(periods=21) * 100
    
    # YoY return (252 trading days)
    df['YoY'] = df['Close'].pct_change(periods=252) * 100
    
    # YTD return
    year_start = df.index.to_series().apply(lambda x: pd.Timestamp(year=x.year, month=1, day=1))
    year_start_prices = df.groupby(year_start)['Close'].transform('first')
    df['YTD'] = ((df['Close'] - year_start_prices) / year_start_prices * 100)
    
    return df

def create_bokeh_chart(commodity_data, commodity_name, output_html):
    """Tạo biểu đồ interactive đẹp với Bokeh"""
    from bokeh.models import ColumnDataSource, CrosshairTool, Range1d
    from bokeh.models.formatters import DatetimeTickFormatter
    
    # Prepare data
    dates = commodity_data.index.to_pydatetime()
    prices = commodity_data['Close'].values
    daily_pct = commodity_data['Daily'].values
    
    # Format dates for display
    date_strings = [d.strftime('%Y-%m-%d') for d in dates]
    
    # Create data source
    source = ColumnDataSource(data={
        'x': dates,
        'y': prices,
        'date_str': date_strings,
        'daily_pct': daily_pct,
        'daily_pct_str': [f"{x:+.2f}%" if pd.notna(x) else "N/A" for x in daily_pct]
    })
    
    # Calculate price range with 5% padding
    price_min = prices.min()
    price_max = prices.max()
    price_range = price_max - price_min
    y_start = price_min - price_range * 0.05
    y_end = price_max + price_range * 0.05
    
    # Create figure với styling đẹp
    p = figure(
        title=f"{commodity_name}",
        x_axis_type='datetime',
        width=1400,
        height=700,
        tools="pan,wheel_zoom,box_zoom,reset,save,crosshair",
        toolbar_location="right",
        sizing_mode='scale_width',
        y_range=Range1d(y_start, y_end)
    )
    
    # Vẽ line chính - KHÔNG CÓ CIRCLE
    line = p.line('x', 'y', source=source, line_width=2.5, color='#2962FF', alpha=0.9)
    
    # Thêm fill area dưới line
    p.varea(x='x', y1=y_start, y2='y', source=source, alpha=0.1, color='#2962FF')
    
    # Hover tool với tooltip đẹp như TradingView
    hover = HoverTool(
        tooltips="""
        <div style="background-color: #1E222D; padding: 12px; border-radius: 4px; border: 1px solid #363A45; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto;">
            <div style="color: #B2B5BE; font-size: 11px; margin-bottom: 6px;">@date_str</div>
            <div style="display: flex; justify-content: space-between; gap: 20px;">
                <div>
                    <div style="color: #787B86; font-size: 11px;">Price</div>
                    <div style="color: #D1D4DC; font-size: 14px; font-weight: 600;">$@y{0,0.00}</div>
                </div>
                <div>
                    <div style="color: #787B86; font-size: 11px;">Daily Change</div>
                    <div style="color: @daily_pct_color; font-size: 14px; font-weight: 600;">@daily_pct_str</div>
                </div>
            </div>
        </div>
        """,
        formatters={'@x': 'datetime'},
        renderers=[line],
        mode='vline',
        line_policy='nearest'
    )
    
    # Add color for daily_pct in tooltip
    source.data['daily_pct_color'] = ['#26A69A' if pd.notna(x) and x >= 0 else '#EF5350' if pd.notna(x) else '#787B86' for x in daily_pct]
    
    p.add_tools(hover)
    
    # Crosshair styling
    crosshair = p.select_one(CrosshairTool)
    crosshair.line_color = '#787B86'
    crosshair.line_alpha = 0.6
    
    # Title styling
    p.title.text_font_size = '18pt'
    p.title.text_color = '#D1D4DC'
    p.title.text_font = 'Helvetica Neue, Arial'
    p.title.align = 'left'
    
    # Axes styling giống TradingView
    p.xaxis.axis_label_text_font_size = '0pt'  # Hide label
    p.yaxis.axis_label = 'Price ($)'
    p.yaxis.axis_label_text_font_size = '11pt'
    p.yaxis.axis_label_text_color = '#787B86'
    p.yaxis.axis_label_standoff = 10
    
    # Format datetime axis
    p.xaxis.formatter = DatetimeTickFormatter(
        hours='%H:%M',
        days='%d %b',
        months='%b %Y',
        years='%Y'
    )
    
    # Grid styling
    p.xgrid.grid_line_color = '#363A45'
    p.xgrid.grid_line_alpha = 0.5
    p.xgrid.grid_line_dash = 'dotted'
    p.ygrid.grid_line_color = '#363A45'
    p.ygrid.grid_line_alpha = 0.5
    p.ygrid.grid_line_dash = 'dotted'
    
    # Background colors - dark theme như TradingView
    p.background_fill_color = '#1E222D'
    p.border_fill_color = '#1E222D'
    p.outline_line_color = '#363A45'
    
    # Axis colors
    p.xaxis.axis_line_color = '#363A45'
    p.yaxis.axis_line_color = '#363A45'
    p.xaxis.major_tick_line_color = '#363A45'
    p.yaxis.major_tick_line_color = '#363A45'
    p.xaxis.minor_tick_line_color = None
    p.yaxis.minor_tick_line_color = None
    p.xaxis.major_label_text_color = '#787B86'
    p.yaxis.major_label_text_color = '#787B86'
    p.xaxis.major_label_text_font_size = '11pt'
    p.yaxis.major_label_text_font_size = '11pt'
    
    # Toolbar styling
    p.toolbar.logo = None  # Remove Bokeh logo
    
    # Save to HTML as a 100% self-contained file
    print(f"    Đang tạo file HTML tự chứa (self-contained) cho: {commodity_name}")
    
    # Dùng file_html và INLINE để nhúng toàn bộ JS/CSS vào file
    html_content = file_html(p, resources=INLINE, title=commodity_name)
    
    # Tự tay ghi nội dung đã nhúng ra file
    with open(output_html, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"    Đã tạo file HTML tự chứa thành công: {output_html}")

    return output_html

def create_commodity_charts(df, 
                            output_file='commodity_charts.xlsx', 
                            period_years=1, 
                            upload_mode=False,
                            local_html_folder='charts_html',
                            github_repo_local_path=None,
                            github_pages_url=None
                           ):
    """
    Vẽ biểu đồ và tạo BẢNG TÓM TẮT cho TẤT CẢ commodities
    vào MỘT sheet duy nhất.
    
    Layout: Chart bên trái (Cột A), Bảng data bên phải (Cột L).
    """
    
    # Đảm bảo date là index
    if 'date' in df.columns:
        df = df.set_index('date')
    
    commodities = df['name'].unique()
    cutoff_date = df.index.max() - pd.DateOffset(years=period_years)
    
    # === THAY ĐỔI 1: TẠO 1 SHEET DUY NHẤT BÊN NGOÀI VÒNG LẶP ===
    wb = Workbook()
    ws = wb.active # Lấy sheet đầu tiên
    ws.title = "Yahoo Finance Summary"
    
    current_row = 1 # Khởi tạo biến đếm hàng

    # Kiểm tra cấu hình dựa trên chế độ (giữ nguyên)
    if upload_mode:
        if not (github_repo_local_path and github_pages_url):
            raise ValueError("LỖI: 'upload_mode=True' nhưng thiếu 'github_repo_local_path' hoặc 'github_pages_url'.")
        if not github_pages_url.endswith('/'):
            github_pages_url += '/'
        print("--- Đang chạy ở chế độ UPLOAD ---")
        os.makedirs(github_repo_local_path, exist_ok=True) 
    else:
        os.makedirs(local_html_folder, exist_ok=True)
        print("--- Đang chạy ở chế độ LOCAL ---")

    for idx, commodity_code in enumerate(commodities):
        commodity_name = COMMODITY_NAMES.get(commodity_code, commodity_code)
        full_name = f"{commodity_name} ({commodity_code})"
        
        print(f"Đang xử lý {full_name}...")
        
        commodity_data_full = df[df['name'] == commodity_code].copy().sort_index()
        commodity_data_full = calculate_returns(commodity_data_full['Close'])
        commodity_data = commodity_data_full[commodity_data_full.index >= cutoff_date].copy()
        
        # === 1. TẠO MATPLOTLIB CHART (Giữ nguyên) ===
        fig, ax = plt.subplots(figsize=(10, 6), dpi=100) # Giảm kích thước ảnh 1 chút
        dates = commodity_data.index
        prices = commodity_data['Close'].values
        ax.plot(dates, prices, color='#3498DB', linewidth=2)
        ax.fill_between(dates, prices, alpha=0.2, color='#3498DB')
        price_min = prices.min(); price_max = prices.max()
        price_range = price_max - price_min
        y_min = price_min - price_range * 0.1; y_max = price_max + price_range * 0.1
        ax.set_ylim(y_min, y_max)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        plt.xticks(rotation=45, ha='right')
        ax.set_title(f'{full_name} - Last {period_years} Year(s)', fontsize=16, pad=10)
        ax.set_ylabel('Close Price ($)', fontsize=12)
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.set_facecolor('#FAFAFA'); fig.patch.set_facecolor('white')
        plt.tight_layout()
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close(fig)
        
        # === 2. TẠO BOKEH INTERACTIVE CHART (VÀ LINK) (Giữ nguyên) ===
        html_filename = f"{commodity_code.replace('=', '_')}.html"
        html_save_path = ""
        excel_hyperlink = ""
        excel_link_text = ""

        if upload_mode:
            html_save_path = os.path.join(github_repo_local_path, html_filename)
            excel_hyperlink = github_pages_url + html_filename
            excel_link_text = "Click to open (GitHub Page)"
        else:
            html_save_path = os.path.join(local_html_folder, html_filename)
            excel_hyperlink = os.path.abspath(html_save_path)
            excel_link_text = "Click to open (Local File)"
            
        create_bokeh_chart(commodity_data, full_name, html_save_path)
        
        # === 3. GHI VÀO EXCEL (LOGIC MỚI) ===
        
        # --- A. Tiêu đề (Gộp A đến T) ---
        ws.merge_cells(f'A{current_row}:T{current_row}')
        cell_title = ws[f'A{current_row}']
        cell_title.value = full_name
        cell_title.font = Font(bold=True, size=16, color='2C3E50')
        cell_title.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[current_row].height = 25
        
        current_row += 1 # Sang hàng mới
        
        # --- B. Link (Cột A) & Stats (Cột L) ---
        # Link
        ws[f'A{current_row}'] = 'Interactive Chart:'
        cell_link = ws[f'B{current_row}']
        cell_link.hyperlink = excel_hyperlink
        cell_link.value = excel_link_text
        cell_link.font = Font(color='0563C1', underline='single')
        cell_link.style = 'Hyperlink'
        
        # Stats
        ws[f'L{current_row}'] = f'Period: Last {period_years} year(s) | Min: ${price_min:,.2f} | Max: ${price_max:,.2f} | Avg: ${prices.mean():,.2f}'
        ws[f'L{current_row}'].font = Font(size=10, color='7F8C8D')
        
        current_row += 2 # Sang hàng mới, chừa 1 hàng trống
        
        # --- C. VỊ TRÍ MỚI: Ảnh (Trái) & Bảng (Phải) ---
        
        # ANCHOR (mỏ neo) cho cả ảnh và bảng
        anchor_row = current_row
        
        # C1. Thêm Ảnh (Bên Trái)
        # Neo ảnh vào cột A
        img = OpenpyxlImage(img_buffer)
        img.width = 600  # 10 * 60 (Rộng 10 cột, từ A-J)
        img.height = 360 # 24 * 15 (Cao 24 hàng)
        ws.add_image(img, f'A{anchor_row}')
        
        # C2. Thêm Bảng (Bên Phải)
        # Bắt đầu bảng từ cột L (cách cột A 11 cột)
        table_start_col = 12 # Cột L

        # Header bảng
        headers = ['Date', 'Close', 'Daily %', 'Weekly %', 'Monthly %', 'YoY %', 'YTD %']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=anchor_row, column=table_start_col + col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data - 10 ngày gần nhất
        recent_data = commodity_data.sort_index(ascending=False).head(10)
        
        for row_idx, (date, row) in enumerate(recent_data.iterrows()):
            data_row = anchor_row + 1 + row_idx # Hàng data (bắt đầu từ hàng_neo + 1)
            
            ws.cell(row=data_row, column=table_start_col, value=date.strftime('%Y-%m-%d'))
            ws.cell(row=data_row, column=table_start_col + 1, value=float(row['Close'])).number_format = '#,##0.00'
            ws.cell(row=data_row, column=table_start_col + 2, value=float(row['Daily']) if pd.notna(row['Daily']) else None).number_format = '0.00'
            ws.cell(row=data_row, column=table_start_col + 3, value=float(row['Weekly']) if pd.notna(row['Weekly']) else None).number_format = '0.00'
            ws.cell(row=data_row, column=table_start_col + 4, value=float(row['Monthly']) if pd.notna(row['Monthly']) else None).number_format = '0.00'
            ws.cell(row=data_row, column=table_start_col + 5, value=float(row['YoY']) if pd.notna(row['YoY']) else None).number_format = '0.00'
            ws.cell(row=data_row, column=table_start_col + 6, value=float(row['YTD']) if pd.notna(row['YTD']) else None).number_format = '0.00'
            
            # Tô màu
            for col in range(2, 7): # Cột Daily -> YTD
                cell = ws.cell(row=data_row, column=table_start_col + col)
                if cell.value and cell.value > 0: cell.font = Font(color='00B050')
                elif cell.value and cell.value < 0: cell.font = Font(color='FF0000')

        # --- D. Cập nhật current_row ---
        # Tăng số hàng bằng chiều cao của ảnh (360px ~ 24 hàng) + 2 hàng đệm
        current_row += 24 + 2 # (360/15 = 24)
        
    # === KẾT THÚC VÒNG LẶP ===
    
    # Điều chỉnh độ rộng cột cho đẹp
    ws.column_dimensions['A'].width = 10 # (Cột A-J là cho ảnh)
    ws.column_dimensions['K'].width = 3  # Cột đệm
    ws.column_dimensions['L'].width = 12 # Cột Date
    ws.column_dimensions['M'].width = 12 # Cột Close
    for col in ['N', 'O', 'P', 'Q', 'R']: # Cột %
        ws.column_dimensions[col].width = 11

    # Save Excel
    wb.save(output_file)
    print(f"\\n✅ Đã xuất thành công file Excel (local): {output_file}")
    print(f"📊 Tổng số commodity: {len(commodities)}")
    print(f"📁 Excel file: {output_file}")