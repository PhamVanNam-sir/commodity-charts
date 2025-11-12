import requests
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import io
import time


# --- BƯỚC 1: Xây dựng bản đồ (map) Tên Commodity -> ID (Giữ nguyên) ---
def get_commodity_map():
    base_url = "https://www.sunsirs.com/uk/"
    page_url = f"{base_url}sectors.html"
    commodity_map = {}
    
    print(f"Đang tải trang danh mục từ {page_url}...")
    try:
        r = requests.get(page_url, headers={'User-Agent': 'Mozilla/5.0'})
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')
        
        link_divs = soup.find_all('div', class_='paddl10')
        
        for div in link_divs:
            links = div.find_all('a')
            for link in links:
                name = link.text.strip()
                href = link.get('href')
                
                if href and 'prodetail-' in href:
                    match = re.search(r'prodetail-(\d+)\.html', href)
                    if match:
                        commodity_id = match.group(1)
                        if name:
                            commodity_map[name] = commodity_id
                            
        print(f"Tìm thấy {len(commodity_map)} commodities.")
        return commodity_map
        
    except Exception as e:
        print(f"LỖI: Không thể lấy dữ liệu commodities: {e}")
        return None

# --- BƯỚC 2 & 3 (Thay đổi hoàn toàn) ---
def create_excel_with_charts(commodity_names_list, output_filename='commodity_charts.xlsx'):
    """
    Hàm chính: Dùng Selenium, screenshot chart,
    CĂN GIỮA TIÊU ĐỀ và LÙI LỀ ẢNH.
    """
    print("Bắt đầu xây dựng bản đồ Tên -> ID...")
    commodity_map = get_commodity_map()
    
    if not commodity_map:
        print("Không thể xây dựng bản đồ. Thoát.")
        return

    # --- Cấu hình Selenium ---
    print("Đang khởi động trình duyệt ảo (Headless Chrome)...")
    # ... (Toàn bộ code Selenium giữ nguyên y hệt) ...
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/5.37.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/5.37.36")
    s = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=s, options=chrome_options)
    print("Trình duyệt ảo đã sẵn sàng.")
    # --------------------------

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commodity Charts"
    
    current_row = 1
    
    scale_factor = 1.3 
    original_width = 550
    original_height = 332
    
    for name_input in commodity_names_list:
        found_name = None
        commodity_id = None
        # ... (Vòng lặp tìm commodity_id giữ nguyên) ...
        for map_name, map_id in commodity_map.items():
            if map_name.lower() == name_input.lower():
                found_name = map_name
                commodity_id = map_id
                break
        
        if not commodity_id:
            print(f"CẢNH BÁO: Không tìm thấy commodity có tên '{name_input}'. Bỏ qua.")
            continue
            
        print(f"Đang xử lý '{found_name}' (ID: {commodity_id})...")
        
        try:
            # 1. & 2. & 3. (Mở trang, tìm
            # ... (Toàn bộ code Selenium để lấy ảnh giữ nguyên) ...
            page_url = f"https://www.sunsirs.com/uk/prodetail-{commodity_id}.html"
            driver.get(page_url)
            time.sleep(2)
            img_xpath = "//img[contains(@src, 'graph.100ppi.com')]"
            img_element = driver.find_element(By.XPATH, img_xpath)
            image_data = img_element.screenshot_as_png
            
            if image_data:
                # 4. Chèn vào Excel (ĐÃ CẬP NHẬT)
                img_file_in_memory = io.BytesIO(image_data)
                
                # --- PHẦN CĂN GIỮA TIÊU ĐỀ ---
                # Gộp 11 cột (A đến W)
                title_cell_start = f'A{current_row}'
                title_cell_end = f'W{current_row}' # Gộp A -> W
                ws.merge_cells(f'{title_cell_start}:{title_cell_end}')
                
                # Lấy ô đã gộp và set giá trị + căn lề
                merged_title_cell = ws[title_cell_start] 
                merged_title_cell.value = found_name
                merged_title_cell.font = Font(bold=True, size=14)
                # Đặt căn lề ngang (horizontal) là 'center'
                merged_title_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[current_row].height = 20 # Tăng chiều cao hàng tiêu đề
                # --- KẾT THÚC PHẦN TIÊU ĐỀ ---

                # --- PHẦN CĂN GIỮA ẢNH ---
                img = Image(img_file_in_memory)
                # Neo ảnh vào cột G (thay vì A) để tạo lề trái
                img_anchor_cell = f'G{current_row + 1}' 
                
                # Scale ảnh
                img.width = original_width * scale_factor
                img.height = original_height * scale_factor
                
                ws.add_image(img, img_anchor_cell)
                # --- KẾT THÚC PHẦN ẢNH ---
                
                # Tăng số hàng (thêm 1 hàng cho tiêu đề)
                rows_to_add = int((img.height / 15) + 3) # +3 để chừa chỗ cho tiêu đề
                current_row += rows_to_add
            
        except Exception as e:
            print(f"LỖI: Không thể chụp ảnh chart cho '{found_name}': {e}")
           
    print(f"\nHoàn tất! Đang lưu file vào {output_filename}...")
    wb.save(output_filename)
    driver.quit()
    print("Đã lưu file thành công.")